#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Conciliação Fiscal de Armazenagem com rastreamento de quantidades"""
import os, sys, pandas as pd, openpyxl as ox, xml.etree.ElementTree as ET, argparse, re
from openpyxl.styles import Font, Alignment, Border, Side

NS = 'http://www.portalfiscal.inf.br/nfe'

def extrair_num(s):
    if pd.isna(s): return None
    m = re.search(r'\d+', str(s).strip())
    return int(m.group()) if m else None

parser = argparse.ArgumentParser()
parser.add_argument('--arquivo', required=True, help='Arquivo Excel com ZSD e razao')
parser.add_argument('--pasta', required=True, help='Pasta com XMLs')
args = parser.parse_args()

ARQUIVO, XML_DIR = args.arquivo, args.pasta

print(f"\n{'='*60}")
print(f"Conciliacao Fiscal de Armazenagem")
print(f"{'='*60}")
print(f"Arquivo: {ARQUIVO}")
print(f"Pasta XMLs: {XML_DIR}\n")

if not os.path.isfile(ARQUIVO):
    print(f"Erro: Arquivo nao encontrado")
    sys.exit(1)
if not os.path.isdir(XML_DIR):
    print(f"Erro: Pasta nao encontrada")
    sys.exit(1)

try:
    df_zsd = pd.read_excel(ARQUIVO, sheet_name='zsd', header=0)
    df_razao = pd.read_excel(ARQUIVO, sheet_name='razao', header=0)
    print(f"[OK] ZSD: {len(df_zsd)} linhas")
    print(f"[OK] Razao: {len(df_razao)} linhas\n")
except Exception as e:
    print(f"Erro ao ler abas: {e}")
    sys.exit(1)

# Colunas do ZSD (0-indexed)
# C=DOC SAP, E=Descricao, P=Nota, V=Chave, X=Material, AB=CFOP, AG=Qtde, AQ=ICMS
ZSD_COL_DOC, ZSD_COL_DESC, ZSD_COL_NF = 2, 4, 15
ZSD_COL_CHAVE, ZSD_COL_MAT, ZSD_COL_CFOP = 21, 23, 27
ZSD_COL_QTD, ZSD_COL_ICMS, ZSD_COL_DATA = 32, 42, 7

# Colunas do Razao (0-indexed)
# G=Referencia, P=Valor, Y=NF RETORNO, Z=DATA RETORNO
RAZAO_COL_REF = 6
RAZAO_COL_NF_RET = 24
RAZAO_COL_DATA_RET = 25

print("Processando ZSD...")
remessas = {}  # (nf, chave) -> {qtd, icms, data}
retornos = []  # [{nf, chave, qtd, icms, data}, ...]

for idx, row in df_zsd.iterrows():
    try:
        doc = str(int(float(str(row.iloc[ZSD_COL_DOC])))) if pd.notna(row.iloc[ZSD_COL_DOC]) else None
        desc = str(row.iloc[ZSD_COL_DESC]).lower() if pd.notna(row.iloc[ZSD_COL_DESC]) else ""
        nf = int(row.iloc[ZSD_COL_NF]) if pd.notna(row.iloc[ZSD_COL_NF]) else None
        chave = str(row.iloc[ZSD_COL_CHAVE]).strip() if pd.notna(row.iloc[ZSD_COL_CHAVE]) else ""
        cfop = str(row.iloc[ZSD_COL_CFOP]).strip() if pd.notna(row.iloc[ZSD_COL_CFOP]) else ""
        qtd = float(row.iloc[ZSD_COL_QTD]) if pd.notna(row.iloc[ZSD_COL_QTD]) else 0.0
        icms = float(row.iloc[ZSD_COL_ICMS]) if pd.notna(row.iloc[ZSD_COL_ICMS]) else 0.0
        data = row.iloc[ZSD_COL_DATA] if pd.notna(row.iloc[ZSD_COL_DATA]) else None

        if not doc or not nf or len(chave) != 44:
            continue

        is_remessa = cfop.startswith(('5', '6')) or 'saida' in desc.replace('í', 'i')
        is_retorno = cfop.startswith(('1', '2', '3')) or 'entrada' in desc

        if is_remessa:
            key = (nf, chave)
            if key not in remessas:
                remessas[key] = {'qtd': 0.0, 'icms': 0.0, 'data': data}
            remessas[key]['qtd'] += qtd
            remessas[key]['icms'] += icms

        elif is_retorno:
            retornos.append({
                'doc': doc, 'nf': nf, 'chave': chave,
                'qtd': qtd, 'icms': icms, 'data': data
            })
    except:
        pass

print(f"  Remessas: {len(remessas)}")
print(f"  Retornos: {len(retornos)}\n")

print("Lendo XMLs...")
xml_roots = {}  # chave44 -> root
xml_refs = {}   # chave44 -> [(nf_remessa, qtd_total), ...]
xml_icms = {}   # chave44 -> icms total

for fname in sorted(os.listdir(XML_DIR)):
    if not fname.endswith('.xml'):
        continue
    try:
        fpath = os.path.join(XML_DIR, fname)
        root = ET.parse(fpath).getroot()

        nnf_el = root.find('.//{%s}nNF' % NS)
        if nnf_el is None:
            continue

        chave44 = fname[:44]
        if len(chave44) != 44:
            continue

        xml_roots[chave44] = root

        # Extrai ICMS total do XML
        vicms_el = root.find('.//{%s}ICMSTot/{%s}vICMS' % (NS, NS))
        icms_xml = 0.0
        if vicms_el is not None and vicms_el.text:
            try:
                icms_xml = float(vicms_el.text)
            except:
                pass
        xml_icms[chave44] = icms_xml

        # PASSO 1: Soma TODAS as quantidades dos itens
        qtd_total = 0.0
        for det in root.findall('.//{%s}det' % NS):
            qCom_el = det.find('.//{%s}qCom' % NS)
            if qCom_el is not None and qCom_el.text:
                try:
                    qtd_total += float(qCom_el.text)
                except:
                    pass

        # PASSO 2: Obtém as referências (C1 dentro de det ou globais)
        nf_qtd_map = {}  # nf -> qtd_total

        # Tenta C1 dentro de cada <det>
        for det in root.findall('.//{%s}det' % NS):
            qCom_el = det.find('.//{%s}qCom' % NS)
            qtd_item = 0.0
            if qCom_el is not None and qCom_el.text:
                try:
                    qtd_item = float(qCom_el.text)
                except:
                    pass

            for el in det.findall('.//{%s}refNFe' % NS):
                chave_ref = (el.text or '').strip()
                if len(chave_ref) == 44:
                    try:
                        nf_ref = int(chave_ref[25:34])
                        if nf_ref not in nf_qtd_map:
                            nf_qtd_map[nf_ref] = 0.0
                        nf_qtd_map[nf_ref] += qtd_item
                    except:
                        pass

        # Se nao encontrou C1 dentro de det, usa refs globais
        if not nf_qtd_map:
            refs_global = []
            for el in root.findall('.//{%s}refNFe' % NS):
                chave_ref = (el.text or '').strip()
                if len(chave_ref) == 44:
                    try:
                        nf_ref = int(chave_ref[25:34])
                        if nf_ref not in refs_global:
                            refs_global.append(nf_ref)
                    except:
                        pass

            # IMPORTANTE: Se múltiplas refs globais, distribui qtd_total
            # Se 1 ref global, toda qtd vai para ela
            if refs_global and qtd_total > 0:
                qtd_por_ref = qtd_total / len(refs_global)
                for nf in refs_global:
                    nf_qtd_map[nf] = qtd_por_ref

        xml_refs[chave44] = list(nf_qtd_map.items())  # [(nf, qtd), ...]
    except:
        pass

print(f"  XMLs processados: {len(xml_roots)}")
print(f"  xml_refs populado com {len(xml_refs)} entradas")

print("\nCruzando retornos com remessas...")
retorno_por_remessa = {}  # (nf_rem, chave_rem) -> [retorno, ...]
retorno_associado = set()  # rastreia quais retornos ja foram associados (evita duplicatas)
xml_processados = set()  # rastreia quais XMLs ja foram processados (evita processar 2x a mesma chave)

for ret in retornos:
    chave_ret = ret['chave']

    # IMPORTANTE: Se o mesmo XML ja foi processado (multiplas linhas ZSD), pula
    if chave_ret in xml_processados:
        continue

    if chave_ret not in xml_roots:
        continue

    xml_processados.add(chave_ret)  # marca como processado

    nfs_ref_qtd = xml_refs.get(chave_ret, [])  # [(nf, qtd), ...]
    if not nfs_ref_qtd:
        continue

    # Um retorno pode referenciar multiplas remessas com quantidades diferentes
    for nf_rem, qtd_xml in nfs_ref_qtd:
        remessa_key = None
        for (nf, chave) in remessas.keys():
            if nf == nf_rem:
                remessa_key = (nf, chave)
                break

        if remessa_key:
            chave_associacao = (chave_ret, remessa_key, qtd_xml)
            if chave_associacao not in retorno_associado:
                if remessa_key not in retorno_por_remessa:
                    retorno_por_remessa[remessa_key] = []

                # Cria novo retorno com quantidade do XML
                ret_xml = {
                    'doc': ret['doc'],
                    'nf': ret['nf'],
                    'chave': ret['chave'],
                    'qtd': qtd_xml,  # usa quantidade do XML, nao do ZSD
                    'icms': ret['icms'],
                    'data': ret['data']
                }
                retorno_por_remessa[remessa_key].append(ret_xml)
                retorno_associado.add(chave_associacao)

print(f"  Associacoes criadas: {len(retorno_por_remessa)}\n")

print("Preenchendo razao...")
wb = ox.load_workbook(ARQUIVO)
ws_razao = wb['razao']

linhas_preenchidas = 0
for idx, row in df_razao.iterrows():
    ref = row.iloc[RAZAO_COL_REF] if pd.notna(row.iloc[RAZAO_COL_REF]) else None
    if not ref:
        continue

    nf_planilha = extrair_num(str(ref))
    if not nf_planilha:
        continue

    remessa_key = None
    for (nf, chave) in remessas.keys():
        if nf == nf_planilha:
            remessa_key = (nf, chave)
            break

    if not remessa_key:
        continue

    if remessa_key not in retorno_por_remessa:
        continue

    ret = retorno_por_remessa[remessa_key][0]
    excel_row = idx + 2

    ws_razao.cell(row=excel_row, column=RAZAO_COL_NF_RET + 1, value=ret['nf'])

    if pd.notna(ret['data']):
        try:
            data_fmt = ret['data'].strftime('%d/%m/%Y') if hasattr(ret['data'], 'strftime') else str(ret['data'])
            ws_razao.cell(row=excel_row, column=RAZAO_COL_DATA_RET + 1, value=data_fmt)
        except:
            pass

    linhas_preenchidas += 1

print(f"  Linhas preenchidas: {linhas_preenchidas}\n")

print("Criando aba 'Controle Fiscal'...")
if 'Controle Fiscal' in wb.sheetnames:
    del wb['Controle Fiscal']

ws_cf = wb.create_sheet('Controle Fiscal', 0)

ws_cf['A1'] = "CONTROLE FISCAL DE ARMAZENAGEM - RESULTADO DA CONCILIACAO"
ws_cf['A1'].font = Font(bold=True, size=12)

ws_cf['A2'] = 'REMESSA'
ws_cf['A2'].font = Font(bold=True)
ws_cf['G2'] = 'RETORNO'
ws_cf['G2'].font = Font(bold=True)
ws_cf['M2'] = 'SALDO'
ws_cf['M2'].font = Font(bold=True)

headers = [
    ('A', 'No Nota Remessa'), ('B', 'Chave (44 dig)'),
    ('C', 'Valor (R$)'), ('D', 'Quantidade'), ('E', 'ICMS (R$)'),
    ('F', 'Data Remessa'),
    ('G', 'No Nota Retorno'), ('H', 'Chave (44 dig)'),
    ('I', 'Valor Ret (R$)'), ('J', 'Qtde Retorn'), ('K', 'ICMS Ret (R$)'),
    ('L', 'Data Retorno'),
    ('M', 'Saldo Valor (R$)'), ('N', 'Saldo Qtde'), ('O', 'Saldo ICMS (R$)')
]

for col, header in headers:
    cell = ws_cf[f'{col}3']
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row_num = 4
fill_roxo = ox.styles.PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid')

for (nf_rem, chave_rem), info_rem in sorted(remessas.items()):
    # Skip se remessa tem ICMS zerado
    if info_rem['icms'] == 0:
        continue

    if (nf_rem, chave_rem) in retorno_por_remessa and retorno_por_remessa[(nf_rem, chave_rem)]:
        rets = retorno_por_remessa[(nf_rem, chave_rem)]

        # Calcula saldo total considerando TODOS os retornos
        saldo_qtd_total = info_rem['qtd']
        saldo_icms_total = info_rem['icms']
        for ret in rets:
            saldo_qtd_total -= ret['qtd']
            saldo_icms_total -= ret['icms']

        # Primeira linha: remessa + primeiro retorno + saldo
        ws_cf[f'A{row_num}'] = nf_rem
        ws_cf[f'B{row_num}'] = chave_rem
        ws_cf[f'C{row_num}'] = 0
        ws_cf[f'D{row_num}'] = info_rem['qtd']
        ws_cf[f'E{row_num}'] = info_rem['icms']

        if info_rem['data']:
            try:
                data_fmt = info_rem['data'].strftime('%d/%m/%Y') if hasattr(info_rem['data'], 'strftime') else str(info_rem['data'])
                ws_cf[f'F{row_num}'] = data_fmt
            except:
                pass

        ret = rets[0]
        ws_cf[f'G{row_num}'] = ret['nf']
        ws_cf[f'H{row_num}'] = ret['chave']
        ws_cf[f'J{row_num}'] = ret['qtd']
        if ret['icms'] != 0:
            ws_cf[f'K{row_num}'] = ret['icms']

        if pd.notna(ret['data']):
            try:
                data_fmt = ret['data'].strftime('%d/%m/%Y') if hasattr(ret['data'], 'strftime') else str(ret['data'])
                ws_cf[f'L{row_num}'] = data_fmt
            except:
                pass

        ws_cf[f'M{row_num}'] = 0
        ws_cf[f'N{row_num}'] = saldo_qtd_total
        ws_cf[f'O{row_num}'] = saldo_icms_total

        row_num += 1

        # Linhas adicionais: APENAS retornos (sem remessa, sem saldo)
        for i in range(1, len(rets)):
            ret = rets[i]
            ws_cf[f'G{row_num}'] = ret['nf']
            ws_cf[f'H{row_num}'] = ret['chave']
            ws_cf[f'J{row_num}'] = ret['qtd']
            if ret['icms'] != 0:
                ws_cf[f'K{row_num}'] = ret['icms']

            if pd.notna(ret['data']):
                try:
                    data_fmt = ret['data'].strftime('%d/%m/%Y') if hasattr(ret['data'], 'strftime') else str(ret['data'])
                    ws_cf[f'L{row_num}'] = data_fmt
                except:
                    pass

            row_num += 1
    else:
        # Remessa sem retorno (e com ICMS != 0)
        ws_cf[f'A{row_num}'] = nf_rem
        ws_cf[f'B{row_num}'] = chave_rem
        ws_cf[f'C{row_num}'] = 0
        ws_cf[f'D{row_num}'] = info_rem['qtd']
        ws_cf[f'E{row_num}'] = info_rem['icms']

        if info_rem['data']:
            try:
                data_fmt = info_rem['data'].strftime('%d/%m/%Y') if hasattr(info_rem['data'], 'strftime') else str(info_rem['data'])
                ws_cf[f'F{row_num}'] = data_fmt
            except:
                pass

        ws_cf[f'M{row_num}'] = 0
        ws_cf[f'N{row_num}'] = info_rem['qtd']
        ws_cf[f'O{row_num}'] = info_rem['icms']
        row_num += 1

# Processa ajustes manuais do razao (retornos que nao vem de XML)
print("Processando ajustes manuais...")
ajustes_encontrados = 0
for idx, row in df_razao.iterrows():
    ref = row.iloc[RAZAO_COL_REF] if pd.notna(row.iloc[RAZAO_COL_REF]) else None
    if not ref: continue
    ref_str = str(ref).upper()

    # Detecta ajuste: contem "AJ" ou referencia nao numerica
    if 'AJ' not in ref_str:
        nf_num = extrair_num(str(ref))
        if nf_num:
            # Ja foi processado como retorno via XML
            if (nf_num, next((c for (n, c) in remessas if n == nf_num), None)) in retorno_por_remessa:
                continue

    # Eh um ajuste manual
    try:
        val = float(row.iloc[15]) if pd.notna(row.iloc[15]) else 0
    except:
        val = 0
    if val >= 0: continue  # apenas retornos (valores negativos)

    # Tenta extrair NF da referencia
    nf_related = extrair_num(str(ref))
    if not nf_related: continue

    remessa_key = None
    for (nf, chave) in remessas.keys():
        if nf == nf_related:
            remessa_key = (nf, chave)
            break

    if not remessa_key: continue

    # Preenche linha de ajuste manual
    nf_rem, chave_rem = remessa_key
    info_rem = remessas[remessa_key]

    ws_cf[f'A{row_num}'] = nf_rem
    ws_cf[f'B{row_num}'] = chave_rem
    ws_cf[f'C{row_num}'] = 0
    ws_cf[f'D{row_num}'] = info_rem['qtd']
    if info_rem['icms'] != 0:
        ws_cf[f'E{row_num}'] = info_rem['icms']

    if info_rem['data']:
        try:
            data_fmt = info_rem['data'].strftime('%d/%m/%Y') if hasattr(info_rem['data'], 'strftime') else str(info_rem['data'])
            ws_cf[f'F{row_num}'] = data_fmt
        except:
            pass

    ws_cf[f'G{row_num}'] = f"{nf_rem} - AJUSTE MANUAL"
    ws_cf[f'J{row_num}'] = 0

    data_lanc = row.iloc[9] if pd.notna(row.iloc[9]) else None
    if data_lanc:
        try:
            data_fmt = data_lanc.strftime('%d/%m/%Y') if hasattr(data_lanc, 'strftime') else str(data_lanc)
            ws_cf[f'L{row_num}'] = data_fmt
        except:
            pass

    # Pinta de roxo
    for col in 'ABCDEFGHIJKLMNO':
        ws_cf[f'{col}{row_num}'].fill = fill_roxo

    row_num += 1
    ajustes_encontrados += 1

if ajustes_encontrados > 0:
    print(f"  Ajustes manuais encontrados: {ajustes_encontrados}")

ws_cf.column_dimensions['B'].width = 48
ws_cf.column_dimensions['H'].width = 48

output = ARQUIVO.replace('.xlsx', '_FISCAL_v10.xlsx')
try:
    wb.save(output)
    print(f"  Arquivo salvo: {output}")
    print(f"  Linhas Controle Fiscal: {row_num - 4}\n")
    print("="*60)
    print("Conciliacao fiscal concluida com sucesso!")
    print("="*60)
except Exception as e:
    print(f"Erro ao salvar: {e}")
    sys.exit(1)
