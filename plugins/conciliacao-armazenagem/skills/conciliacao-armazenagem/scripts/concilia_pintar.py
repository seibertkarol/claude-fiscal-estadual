import pandas as pd
import xml.etree.ElementTree as ET
import os, re, sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

NS = 'http://www.portalfiscal.inf.br/nfe'

# ---------------------------------------------------------------
# PARAMETROS — cada fazenda/colega informa o proprio arquivo e pasta de XML
# Uso: py -3.14 concilia_pintar.py --arquivo="EXPORT_XXXX.xlsx" --pasta="INNFE_XXXX"
# Pressupoe que TODAS as planilhas seguem o mesmo layout de colunas (Planilha6,
# zsd, mesma posicao das colunas Referencia/Data lancamento/Valor/etc).
# ---------------------------------------------------------------
ARQUIVO = None
XML_DIR = None
for arg in sys.argv[1:]:
    if arg.startswith('--arquivo='):
        ARQUIVO = arg.split('=', 1)[-1].strip().strip('"')
    elif arg.startswith('--pasta='):
        XML_DIR = arg.split('=', 1)[-1].strip().strip('"')

if not ARQUIVO:
    print("ERRO: informe o arquivo Excel com --arquivo=\"NOME.xlsx\"")
    sys.exit(1)
if not XML_DIR:
    print("ERRO: informe a pasta de XMLs com --pasta=\"NOME_DA_PASTA\"")
    sys.exit(1)

FILL_ROSA    = PatternFill('solid', fgColor='F2CEEF')   # ZEROU total
FILL_LARANJA = PatternFill('solid', fgColor='E97132')   # PARCIAL
FILL_AZUL    = PatternFill('solid', fgColor='467886')   # SEM REMESSA
FILL_AMARELO = PatternFill('solid', fgColor='FFD700')   # DIVERGENTE: ICMS do XML != valor lancado no SAP
FONT_BRANCO  = Font(color='FFFFFF')

# ---------------------------------------------------------------
# DETECCAO DE ESTADO: a linha 1 da Planilha6 e SEMPRE a linha de total
# (formula na coluna de valor), igual em todas as planilhas da usuaria —
# isso NAO e algo que so meu script cria, e o padrao dela. O que varia e:
#   1. Se a coluna J (DATA DO RETORNO) ja foi inserida numa rodagem anterior
#   2. Se a linha de total ja tem o texto "TOTAL FILTRADO" (escrito por mim)
# Detectamos cada coisa separadamente, olhando o conteudo real das celulas,
# em vez de depender de um texto exato que pode nao estar la.
# ---------------------------------------------------------------
from openpyxl import load_workbook as _peek_wb
_wb_peek = _peek_wb(ARQUIVO, read_only=True, data_only=True)

def _resolver_aba(nome_esperado):
    """Acha o nome real da aba ignorando maiusculas/minusculas
    (ex: planilha pode ter 'ZSD' ou 'zsd')."""
    for nome in _wb_peek.sheetnames:
        if nome.strip().lower() == nome_esperado.lower():
            return nome
    raise ValueError(
        f"Aba '{nome_esperado}' nao encontrada. Abas disponiveis: {_wb_peek.sheetnames}"
    )

try:
    ABA_PLANILHA6 = _resolver_aba('Planilha6')
except ValueError:
    ABA_PLANILHA6 = _resolver_aba('razao')
ABA_ZSD       = _resolver_aba('zsd')

_ws_peek = _wb_peek[ABA_PLANILHA6]

# A coluna G (Referencia) tem texto em toda linha de dados. Se a linha 2
# tiver esse texto e a linha 1 nao, a linha 1 e a linha de total/cabecalho
# vazio — ou seja, o cabeçalho de verdade esta na linha 2.
_val_linha1_G = _ws_peek.cell(row=1, column=7).value
_val_linha2_G = _ws_peek.cell(row=2, column=7).value
LINHA1_E_TOTAL = (
    (_val_linha1_G in (None, '')) and isinstance(_val_linha2_G, str) and _val_linha2_G.strip() != ''
)

# A coluna J (10) ja foi inserida (DATA DO RETORNO) numa rodagem anterior?
_linha_cabecalho_1based = 2 if LINHA1_E_TOTAL else 1
_val_col_j = _ws_peek.cell(row=_linha_cabecalho_1based, column=10).value
# Aceita variacoes de texto ("DATA DO RETORNO", "DATA DE RETORNO", etc) —
# o importante e que a celula comece com "DATA" e contenha "RETORNO", ja
# que diferentes planilhas/usuarias podem ter digitado isso de forma levemente
# diferente.
COLUNA_J_JA_EXISTE = (
    isinstance(_val_col_j, str)
    and 'DATA' in _val_col_j.strip().upper()
    and 'RETORNO' in _val_col_j.strip().upper()
)

_wb_peek.close()

PANDAS_HEADER = 1 if LINHA1_E_TOTAL else 0   # linha do cabecalho real (0-based)
COL_OFFSET    = 1 if COLUNA_J_JA_EXISTE else 0   # deslocamento das colunas J em diante

# Mantem o nome JA_PROCESSADO para o restante do script (controla insercoes)
JA_PROCESSADO_LINHA = LINHA1_E_TOTAL
JA_PROCESSADO_COLUNA = COLUNA_J_JA_EXISTE

USANDO_RAZAO = ABA_PLANILHA6.strip().lower() == 'razao'

COL_REFERENCIA   = 6                  # nunca desloca (fica antes da coluna I)
if USANDO_RAZAO:
    COL_DATA_LANC = 9    # col J no razao (Data de lancamento)
    COL_VALOR     = 15   # col P no razao
else:
    COL_DATA_LANC = 10 + COL_OFFSET
    COL_VALOR     = 17 + COL_OFFSET

print(f"Linha 1 {'JA E a linha de total' if LINHA1_E_TOTAL else 'NAO existe ainda (sera criada)'} | "
      f"Coluna DATA DO RETORNO {'JA EXISTE' if COLUNA_J_JA_EXISTE else 'NAO existe ainda (sera criada)'}.")

print("Carregando dados...")
df_p6  = pd.read_excel(ARQUIVO, sheet_name=ABA_PLANILHA6, header=PANDAS_HEADER)
df_zsd = pd.read_excel(ARQUIVO, sheet_name=ABA_ZSD,       header=0)

df_p6['_valor'] = pd.to_numeric(df_p6.iloc[:, COL_VALOR], errors='coerce')

COL_NOTA_RET = 24 if USANDO_RAZAO else 8  # col Y no razao, col I no padrao
print(f"Coluna NOTA DE RETORNO: '{df_p6.columns[COL_NOTA_RET]}'")

# ---------------------------------------------------------------
# MAPA 1: ZSD doc_sap -> nfe
# ---------------------------------------------------------------
zsd_by_doc = {}
for _, row in df_zsd.iterrows():
    doc = str(int(row.iloc[4])) if pd.notna(row.iloc[4]) else None
    if doc and doc not in zsd_by_doc:
        zsd_by_doc[doc] = {'nfe': row.iloc[50]}

# ---------------------------------------------------------------
# MAPA 2: P6 retorno nfe -> indices + valor
# ---------------------------------------------------------------
def ref_to_str(v):
    try:    return str(int(float(v)))
    except: return None

p6_retorno_by_nfe = {}
p6_retorno_valor  = {}
for idx, row in df_p6.iterrows():
    val = row['_valor']
    if pd.isna(val) or val >= 0: continue
    ref = ref_to_str(row.iloc[6])
    if not ref: continue
    info = zsd_by_doc.get(ref)
    if not info: continue
    try: nfe = int(info['nfe'])
    except: continue
    p6_retorno_by_nfe.setdefault(nfe, []).append(idx)
    p6_retorno_valor[nfe] = p6_retorno_valor.get(nfe, 0.0) + float(val)

print(f"Retornos identificados via ZSD: {len(p6_retorno_by_nfe)}")

# Data de lancamento de cada retorno (coluna K = idx 10), para preencher
# na coluna J (DATA DO RETORNO) de toda linha que ele referencia
nnf_data_lancamento = {}
for nfe, idxs in p6_retorno_by_nfe.items():
    for idx in idxs:
        data_lanc = df_p6.iloc[idx, COL_DATA_LANC]
        if pd.notna(data_lanc):
            nnf_data_lancamento[nfe] = data_lanc
            break

# ---------------------------------------------------------------
# MAPA 3: P6 remessa nf_num -> indices + valor original
# ---------------------------------------------------------------
def extrair_num(v):
    s = str(v).strip()
    if s.lower() in ('nan','none',''): return None
    if '-' in s: s = s.split('-')[0]
    s = s.lstrip('0')
    try:    return int(s)
    except: return None

p6_remessa_by_num = {}   # nf_num -> [pandas_idx]
p6_remessa_valor  = {}   # nf_num -> valor total original
for idx, row in df_p6.iterrows():
    val = row['_valor']
    if pd.isna(val) or val <= 0: continue
    n = extrair_num(str(row.iloc[6]))
    if n is not None:
        p6_remessa_by_num.setdefault(n, []).append(idx)
        p6_remessa_valor[n] = p6_remessa_valor.get(n, 0.0) + float(val)

# Saldo disponivel de cada remessa (vai sendo consumido conforme retornos processam)
saldo_remessa = dict(p6_remessa_valor)

# Rastreia quais retornos C1 consumiram saldo de cada remessa (por nf_num)
# Somente esses definem se a remessa fica rosa ou laranja
from collections import defaultdict
saldo_consumers = defaultdict(list)   # nf_num -> [nnf_retorno, ...]

# ---------------------------------------------------------------
# EXTRACAO DE NUMEROS DE TEXTO (C2/C3)
# ---------------------------------------------------------------
EXPORT_CTX = ['NFE DE EXPORTACAO','NF DE EXPORTACAO','GUIA DE SAIDA',
              'NAVIO','CONTAINER','BOOKING','CROSSDOCKING','DESTINO:','RES:']
PREFIXOS   = ['NOTAS FISCAIS','NOTA FISCAL','NOTAS','NOTA',
              'PESOS:','PESO:','NFE','NF','N:','N ']

def extrair_nums_texto(txt):
    """
    Extrai numeros de NF referenciados em texto livre (infCpl/xTexto).
    Captura LISTAS de numeros separados por virgula e/ou "E"
    (ex: "NOTAS FISCAIS 66377, 66384 E 66389" -> 66377, 66384, 66389),
    nao apenas o primeiro numero apos o prefixo.
    """
    if not txt: return set()
    nums, tu = set(), txt.upper()
    for m in re.finditer(r'(\d{5,9})-00[123]', txt):
        nums.add(int(m.group(1)))
    for pref in PREFIXOS:
        pos = 0
        while True:
            i2 = tu.find(pref, pos)
            if i2 == -1: break
            sub = txt[i2+len(pref):i2+len(pref)+150]
            # Casa uma sequencia: primeiro numero, depois repeticoes de
            # (virgula/"E"/"e") + numero, parando no primeiro token que nao bate
            m = re.match(
                r'\s*[:\-]?\s*(\d{5,9}(?:\s*(?:,|/|\bE\b)\s*\d{5,9})*)',
                sub, re.IGNORECASE
            )
            if m:
                ctx = tu[max(0, i2-60):i2]
                if not any(c in ctx for c in EXPORT_CTX):
                    for numstr in re.findall(r'\d{5,9}', m.group(1)):
                        nums.add(int(numstr))
            pos = i2 + len(pref)
    return nums

def _ordenar_refs_para_alocacao(refs_nums):
    """
    Processa remessas EXCLUSIVAS primeiro, e as COMPARTILHADAS (referenciadas
    por mais de um retorno) por ultimo. Isso evita que um retorno monopolize
    o saldo inteiro de uma remessa compartilhada so porque foi processado
    primeiro (por ordem de NF) — ele so consome da remessa compartilhada o
    que sobrar depois de cobrir tudo que puder com remessas exclusivas,
    deixando o restante correto disponivel para o outro retorno que tambem
    precisa dela.
    """
    return sorted(refs_nums, key=lambda n: 1 if n in compartilhadas else 0)

def buscar_remessas(refs_nums, val_retorno_abs):
    """
    Busca remessas com saldo disponivel.
    Aloca apenas o minimo entre saldo e o que o retorno ainda precisa.
    Retorna (indices_p6, soma_alocada, refs_encontradas).
    """
    idxs, soma, vistos, refs_found = [], 0.0, set(), set()
    restante = val_retorno_abs
    for n in _ordenar_refs_para_alocacao(refs_nums):
        saldo = saldo_remessa.get(n, 0.0)
        if saldo <= 0: continue
        for idx in p6_remessa_by_num.get(n, []):
            if idx in vistos: continue
            v = df_p6.at[idx, '_valor']
            if pd.notna(v) and v > 0:
                alocado = min(saldo, restante)
                idxs.append(idx)
                soma += alocado
                restante -= alocado
                vistos.add(idx)
                refs_found.add(n)
    return idxs, soma, refs_found

def consumir_saldo(refs_nums, val_retorno_abs, nnf_retorno):
    """Desconta o saldo e registra quem consumiu cada remessa.
    Usa a mesma ordem (exclusivas primeiro) que buscar_remessas, para o
    valor efetivamente descontado ficar consistente com o que foi casado."""
    refs_nums = _ordenar_refs_para_alocacao(refs_nums)
    restante = val_retorno_abs
    for n in refs_nums:
        saldo = saldo_remessa.get(n, 0.0)
        if saldo <= 0: continue
        usado = min(saldo, restante)
        saldo_remessa[n] = max(0.0, saldo - usado)
        if nnf_retorno not in saldo_consumers[n]:
            saldo_consumers[n].append(nnf_retorno)
        restante -= usado
        if restante <= 0: break

# ---------------------------------------------------------------
# PRE-PROCESSAMENTO: le todos os XMLs
# ---------------------------------------------------------------
refs_por_retorno = {}
xml_roots = {}
nnf_vicms_xml = {}  # nnf -> vICMS declarado no XML (para validar contra o lancado no SAP)
for fname in sorted(os.listdir(XML_DIR)):
    if not fname.endswith('.xml'): continue
    root = ET.parse(os.path.join(XML_DIR, fname)).getroot()
    nnf_el = root.find('.//{%s}nNF' % NS)
    if nnf_el is None: continue
    nnf = int(nnf_el.text)
    refs = set()
    for el in root.findall('.//{%s}refNFe' % NS):
        chave = (el.text or '').strip()
        if len(chave) == 44:
            try: refs.add(int(chave[25:34]))
            except: pass
    refs_por_retorno[nnf] = refs
    xml_roots[nnf] = root

    vicms_el = root.find('.//{%s}ICMSTot/{%s}vICMS' % (NS, NS))
    if vicms_el is not None and vicms_el.text:
        try: nnf_vicms_xml[nnf] = float(vicms_el.text)
        except: pass

# Detecta remessas compartilhadas
retornos_por_remessa = {}
for nnf, refs in refs_por_retorno.items():
    for r in refs:
        retornos_por_remessa.setdefault(r, set()).add(nnf)
compartilhadas = {r: nfs for r, nfs in retornos_por_remessa.items() if len(nfs) > 1}
if compartilhadas:
    print(f"Remessas compartilhadas (retorno parcial): {len(compartilhadas)}")

# ---------------------------------------------------------------
# PASSO 1: CONCILIACAO — processa em ordem crescente de NF
# determina status de cada retorno e quais remessas ele consumiu
# ---------------------------------------------------------------
print("\nPasso 1: Processando XMLs (calculando status)...")

xmls_ordenados = sorted(xml_roots.keys())

status_retorno   = {}  # nnf -> 'ZEROU' | 'PARCIAL' | 'SEM_REMESSA' | 'SEM_RAZAO'
ret_rem_idxs     = {}  # nnf -> [pandas_idx das remessas consumidas]
ret_refs_usadas  = {}  # nnf -> set de nf_nums de remessas consumidas
nota_ret_por_rem = {}  # pandas_idx_remessa -> [nnf_retorno, ...]
log_result = []
relatorio_rows = []  # dados estruturados de cada retorno, para o relatorio laranja
cnt = {'ZEROU': 0, 'PARCIAL': 0, 'SEM_REMESSA': 0, 'SEM_RAZAO': 0}

for nnf in xmls_ordenados:
    root = xml_roots[nnf]
    ret_idxs = p6_retorno_by_nfe.get(nnf, [])
    val_ret  = p6_retorno_valor.get(nnf, 0.0)
    val_ret_abs = abs(val_ret)

    if not ret_idxs:
        status_retorno[nnf] = 'SEM_RAZAO'
        cnt['SEM_RAZAO'] += 1
        log_result.append(f'[?] NF {nnf:>6} | SEM_RETORNO_P6')
        relatorio_rows.append({
            'nnf': nnf, 'status': 'SEM_RETORNO_P6', 'criterio': '-',
            'n_remessas': 0, 'val_retorno': 0.0, 'val_remessas': 0.0,
            'diferenca': 0.0, 'referencias': '',
            'vicms_xml': nnf_vicms_xml.get(nnf), 'dif_icms': None,
        })
        continue

    # C1: refNFe
    refs_c1  = refs_por_retorno.get(nnf, set())
    rem_idxs, val_rem, refs_found = [], 0.0, set()
    criterio = ''

    if refs_c1:
        rem_idxs, val_rem, refs_found = buscar_remessas(refs_c1, val_ret_abs)
        if rem_idxs: criterio = 'C1:refNFe'

    # C2: infAdProd + xPed
    if not rem_idxs:
        refs_c2 = set()
        for tag in ['infAdProd','xPed']:
            for el in root.findall('.//{%s}%s' % (NS, tag)):
                refs_c2 |= extrair_nums_texto(el.text or '')
        if refs_c2:
            rem_idxs, val_rem, refs_found = buscar_remessas(refs_c2, val_ret_abs)
            if rem_idxs: criterio = 'C2:infAdProd+xPed'

    # C3: infCpl + xTexto
    if not rem_idxs:
        refs_c3 = set()
        for tag in ['infCpl','xTexto']:
            for el in root.findall('.//{%s}%s' % (NS, tag)):
                refs_c3 |= extrair_nums_texto(el.text or '')
        if refs_c3:
            rem_idxs, val_rem, refs_found = buscar_remessas(refs_c3, val_ret_abs)
            if rem_idxs: criterio = 'C3:infCpl+xTexto'

    # Classifica status
    if not rem_idxs:
        status = 'SEM_REMESSA'; cnt['SEM_REMESSA'] += 1
    else:
        dif = val_ret + val_rem
        if abs(dif) <= 0.05:
            status = 'ZEROU';   cnt['ZEROU'] += 1
        else:
            status = 'PARCIAL'; cnt['PARCIAL'] += 1

    # VERIFICACAO DE INTEGRIDADE: o valor de ICMS declarado no XML deve bater
    # com o valor lancado no SAP/Planilha6 para este retorno. Se houver uma
    # divergencia (lancamento errado, corrigido depois por uma linha de
    # ajuste separada), NAO deixamos pintar de rosa mesmo que a soma feche —
    # marcamos como DIVERGENTE_ICMS (cor amarela) para revisao manual.
    TOLERANCIA_ICMS = 0.10  # ate 10 centavos de diferenca e considerado igual
    vicms_xml = nnf_vicms_xml.get(nnf)
    if vicms_xml is not None and abs(vicms_xml - val_ret_abs) > TOLERANCIA_ICMS:
        status = 'DIVERGENTE_ICMS'
        cnt.setdefault('DIVERGENTE_ICMS', 0)
        cnt['DIVERGENTE_ICMS'] += 1
        # Desfaz a contagem que foi feita acima antes de sabermos da divergencia
        if not rem_idxs: cnt['SEM_REMESSA'] -= 1
        elif abs(val_ret + val_rem) <= 0.05: cnt['ZEROU'] -= 1
        else: cnt['PARCIAL'] -= 1

    status_retorno[nnf] = status

    # Consome saldo (e registra como consumidor, elegivel a rosa) quando:
    #   - e C1 (chave eletronica, sempre confiavel, mesmo se PARCIAL)
    #   - OU e C2/C3 (texto) MAS zerou exato — so confiamos em C2/C3 quando
    #     o valor bate 100%, para nao contaminar com falso positivo parcial
    #   - E NAO tem divergencia de ICMS (lancamento suspeito nao consome saldo)
    if rem_idxs and status != 'DIVERGENTE_ICMS' and (criterio == 'C1:refNFe' or status == 'ZEROU'):
        consumir_saldo(refs_found, val_ret_abs, nnf)

    # Registra quais remessas este retorno consumiu
    ret_rem_idxs[nnf]   = rem_idxs
    ret_refs_usadas[nnf] = refs_found

    # Mapeia: remessa_idx -> quais retornos a usaram
    for idx in rem_idxs:
        nota_ret_por_rem.setdefault(idx, [])
        if nnf not in nota_ret_por_rem[idx]:
            nota_ret_por_rem[idx].append(nnf)

    dif_txt = f'dif=R${val_ret+val_rem:,.2f}' if rem_idxs else ''
    ic = '[OK]' if status=='ZEROU' else ('[~]' if status=='PARCIAL' else '[X]')
    log_result.append(
        f'{ic} NF {nnf:>6} | {status:<12} | {criterio:<22} '
        f'| ret={len(ret_idxs)} rem={len(rem_idxs)} '
        f'| R${val_ret:,.2f} + R${val_rem:,.2f} {dif_txt}'
    )

    # Guarda dados estruturados para o relatorio (usado no relatorio laranja)
    relatorio_rows.append({
        'nnf': nnf,
        'status': status,
        'criterio': criterio or '-',
        'n_remessas': len(rem_idxs),
        'val_retorno': val_ret,
        'val_remessas': val_rem,
        'diferenca': val_ret + val_rem,
        'referencias': ', '.join(str(r) for r in sorted(refs_found)) if rem_idxs else '',
        'vicms_xml': vicms_xml,
        'dif_icms': (round(vicms_xml - val_ret_abs, 2) if vicms_xml is not None else None),
    })

# ---------------------------------------------------------------
# PASSO 2+3: DETERMINA COR — ALGORITMO ITERATIVO (ponto fixo)
#
# Regra do ROSA:
#   Um grupo de retornos + remessas fica ROSA quando forma um ciclo
#   completamente fechado: todos os retornos do grupo sao ZEROU,
#   todas as remessas do grupo tem saldo = 0.
#   Ex: NF 722 + NF 756 compartilham remessa 68582. Juntos zeram -> todos ROSA.
#
# Garantia matematica: filtro rosa = soma R$0.
# O loop converge porque cores so podem ser perdidas (nunca ganhas).
# ---------------------------------------------------------------
print("Passo 2/3: Algoritmo iterativo de cores (ponto fixo)...")

FILL_MAP = {'rosa': FILL_ROSA, 'laranja': FILL_LARANJA, 'azul': FILL_AZUL, 'amarelo': FILL_AMARELO}
pintura_p6 = {}   # pandas_idx -> (cor, nnf, criterio)

# Estado inicial de cor dos retornos (pelo status da conciliacao)
cor_retorno = {}   # nnf -> 'rosa'|'laranja'|'azul'|'amarelo'
for nnf in xmls_ordenados:
    st = status_retorno.get(nnf)
    if st in ('SEM_RAZAO', None): continue
    cor_retorno[nnf] = {
        'ZEROU': 'rosa', 'PARCIAL': 'laranja', 'SEM_REMESSA': 'azul',
        'DIVERGENTE_ICMS': 'amarelo',
    }.get(st, 'laranja')

# Registra retornos nas linhas de retorno da Planilha6
for nnf in xmls_ordenados:
    if nnf not in cor_retorno: continue
    for idx in p6_retorno_by_nfe.get(nnf, []):
        nota_ret_por_rem.setdefault(idx, [])
        if nnf not in nota_ret_por_rem[idx]:
            nota_ret_por_rem[idx].append(nnf)

MAX_ITER = 20
for iteracao in range(1, MAX_ITER + 1):
    mudancas = 0

    # --- Determina cor das REMESSAS ---
    cor_remessa = {}   # pandas_idx -> 'rosa'|'laranja'
    for idx, nfs_list in nota_ret_por_rem.items():
        val = df_p6.at[idx, '_valor']
        if pd.isna(val) or val <= 0:
            continue  # linha de retorno

        nf_num = extrair_num(str(df_p6.iloc[idx, 6]))
        retornos_que_usaram = [n for n in nfs_list if n in ret_rem_idxs and idx in ret_rem_idxs[n]]
        if not retornos_que_usaram:
            continue

        saldo_final = saldo_remessa.get(nf_num, p6_remessa_valor.get(nf_num, 1.0))
        saldo_zerado = saldo_final <= 0.05

        # Todos os consumidores C1 precisam ser atualmente rosa
        consumers_c1 = saldo_consumers.get(nf_num, [])
        todos_rosa = bool(consumers_c1) and all(cor_retorno.get(n) == 'rosa' for n in consumers_c1)

        cor_remessa[idx] = 'rosa' if (saldo_zerado and todos_rosa) else 'laranja'

    # --- Verifica e ajusta cor dos RETORNOS ZEROU ---
    for nnf in xmls_ordenados:
        if cor_retorno.get(nnf) != 'rosa': continue
        rem_idxs_nf = ret_rem_idxs.get(nnf, [])
        alguma_nao_rosa = any(
            cor_remessa.get(idx, 'laranja') != 'rosa'
            for idx in rem_idxs_nf
            if pd.notna(df_p6.at[idx, '_valor']) and df_p6.at[idx, '_valor'] > 0
        )
        if alguma_nao_rosa:
            cor_retorno[nnf] = 'laranja'
            mudancas += 1

    print(f"  iteracao {iteracao}: {mudancas} retornos ajustados")
    if mudancas == 0:
        print(f"  Ponto fixo atingido em {iteracao} iteracao(oes).")
        break

# Remessas referenciadas por um retorno DIVERGENTE_ICMS tambem ficam amarelas
# (a nao ser que ja estejam legitimamente rosa por outro consumidor real),
# para que a usuaria veja o par inteiro (retorno + remessas) junto na revisao.
qtd_remessas_amarelas = 0
for nnf, cor in cor_retorno.items():
    if cor != 'amarelo': continue
    for idx in ret_rem_idxs.get(nnf, []):
        if cor_remessa.get(idx) != 'rosa':
            if cor_remessa.get(idx) != 'amarelo':
                qtd_remessas_amarelas += 1
            cor_remessa[idx] = 'amarelo'
if qtd_remessas_amarelas:
    print(f"  {qtd_remessas_amarelas} remessas marcadas em amarelo (referenciadas por retorno divergente).")

# Constroi pintura_p6 final
for nnf, cor in cor_retorno.items():
    for idx in p6_retorno_by_nfe.get(nnf, []):
        pintura_p6[idx] = (cor, nnf, '')

for idx, cor in cor_remessa.items():
    nfs_list = nota_ret_por_rem.get(idx, [])
    retornos_que_usaram = [n for n in nfs_list if n in ret_rem_idxs and idx in ret_rem_idxs[n]]
    ref = retornos_que_usaram[0] if retornos_que_usaram else 0
    pintura_p6[idx] = (cor, ref, '')

for idx, cor in cor_remessa.items():
    nfs_list = nota_ret_por_rem.get(idx, [])
    retornos_que_usaram = [n for n in nfs_list if n in ret_rem_idxs and idx in ret_rem_idxs[n]]
    ref = retornos_que_usaram[0] if retornos_que_usaram else 0
    pintura_p6[idx] = (cor, ref, '')

# ---------------------------------------------------------------
# APLICA PINTURA NA PLANILHA6
# ---------------------------------------------------------------
print(f"Aplicando cores em {len(pintura_p6)} linhas da Planilha6...")

wb = load_workbook(ARQUIVO)
ws = wb[ABA_PLANILHA6]

# Insere a coluna J (DATA DO RETORNO) e a linha de total SOMENTE se ainda
# nao existirem — e ANTES de pintar, para que os calculos de linha/coluna
# usados na pintura ja reflitam o layout final. As duas coisas sao
# verificadas de forma INDEPENDENTE (a planilha pode ja ter uma linha de
# total no padrao da usuaria sem ainda ter a coluna J, por exemplo).
if not JA_PROCESSADO_COLUNA:
    ws.insert_cols(10)
    _header_row_atual = 2 if JA_PROCESSADO_LINHA else 1
    ws.cell(row=_header_row_atual, column=10, value='DATA DO RETORNO')

if not JA_PROCESSADO_LINHA:
    ws.insert_rows(1)

# O resultado final SEMPRE tem 1 linha TOTAL FILTRADO no topo (inserida agora
# ou ja existente de antes), entao a linha excel final = pandas_idx + 3, sempre.
EXCEL_ROW_BASE = 3

for pandas_idx, (cor, nnf, criterio) in pintura_p6.items():
    excel_row = pandas_idx + EXCEL_ROW_BASE
    fill = FILL_MAP[cor]
    for cell in ws[excel_row]:
        cell.fill = fill
        cell.font = FONT_BRANCO if cor == 'azul' else Font()

# Preenche coluna I com NF(s) de retorno (remessas e retornos)
COL_I_EXCEL = 9
print(f"Preenchendo NOTA DE RETORNO em {len(nota_ret_por_rem)} linhas...")
COL_J_EXCEL = 10

def fmt_data(d):
    try: return d.strftime('%d/%m/%Y')
    except: return str(d)

for pandas_idx, nfs_list in nota_ret_por_rem.items():
    if pandas_idx not in pintura_p6: continue
    excel_row = pandas_idx + EXCEL_ROW_BASE
    valor_col_i = ', '.join(str(n) for n in sorted(set(nfs_list)))
    ws.cell(row=excel_row, column=COL_I_EXCEL, value=valor_col_i)

    # Coluna J (DATA DO RETORNO): data de lancamento da(s) nota(s) de retorno
    datas_distintas = []
    for n in sorted(set(nfs_list)):
        d = nnf_data_lancamento.get(n)
        if d is not None and d not in datas_distintas:
            datas_distintas.append(d)

    if len(datas_distintas) == 1:
        # Uma so data -> escreve como data real (mantem formatacao de data no Excel)
        cell_j = ws.cell(row=excel_row, column=COL_J_EXCEL, value=datas_distintas[0])
        cell_j.number_format = 'DD/MM/YYYY'
    elif len(datas_distintas) > 1:
        # Datas diferentes (retornos parciais em datas distintas) -> texto separado por virgula
        ws.cell(row=excel_row, column=COL_J_EXCEL, value=', '.join(fmt_data(d) for d in datas_distintas))

# Linha de SUBTOTAL no topo (a linha em si ja foi inserida mais acima,
# antes da pintura — aqui so escrevemos a formula e o estilo)
ws['A1'] = 'TOTAL FILTRADO'

from openpyxl.utils import get_column_letter
col_valor_letra = get_column_letter(COL_VALOR + 1 + (0 if JA_PROCESSADO_COLUNA else 1))
ultima_linha = ws.max_row
formula_cell = f'{col_valor_letra}1'
ws[formula_cell] = f'=SUBTOTAL(9,{col_valor_letra}3:{col_valor_letra}{ultima_linha})'
ws[formula_cell].number_format = '#,##0.00'
fill_total = PatternFill('solid', fgColor='1F4E79')
font_branco_bold = Font(bold=True, color='FFFFFF')
for col in range(1, ws.max_column + 1):
    ws.cell(row=1, column=col).fill = fill_total
    ws.cell(row=1, column=col).font = font_branco_bold
ws[formula_cell].number_format = '#,##0.00'

# Conta remessas/retornos por cor (precisa ser antes do save, para escrever na aba Resumo)
cnt_rem = {'rosa': 0, 'laranja': 0, 'azul': 0, 'amarelo': 0}
cnt_ret = {'rosa': 0, 'laranja': 0, 'azul': 0, 'amarelo': 0}
for pandas_idx, (cor, nnf, _) in pintura_p6.items():
    val = df_p6.at[pandas_idx, '_valor']
    if pd.notna(val):
        if val > 0: cnt_rem[cor] = cnt_rem.get(cor,0) + 1
        else:       cnt_ret[cor] = cnt_ret.get(cor,0) + 1

# Soma monetaria por cor (mesma logica do diagnostico mais abaixo)
soma_rosa_rem = soma_rosa_ret = soma_laranja_rem = soma_laranja_ret = 0.0
soma_amarelo_rem = soma_amarelo_ret = 0.0
for pandas_idx, (cor, nnf, _) in pintura_p6.items():
    val = df_p6.at[pandas_idx, '_valor']
    if pd.isna(val): continue
    if cor == 'rosa':
        if val > 0: soma_rosa_rem += val
        else:       soma_rosa_ret += val
    elif cor == 'laranja':
        if val > 0: soma_laranja_rem += val
        else:       soma_laranja_ret += val
    elif cor == 'amarelo':
        if val > 0: soma_amarelo_rem += val
        else:       soma_amarelo_ret += val

# ---------------------------------------------------------------
# ABA "RESUMO" — contagem desta rodagem (igual ao VBA antigo)
# ---------------------------------------------------------------
NOME_ABA_RESUMO = 'Resumo Conciliacao'
if NOME_ABA_RESUMO in wb.sheetnames:
    del wb[NOME_ABA_RESUMO]
ws_resumo = wb.create_sheet(NOME_ABA_RESUMO, 0)  # insere como primeira aba

fill_header = PatternFill('solid', fgColor='1F4E79')
font_header = Font(bold=True, color='FFFFFF', size=12)
fill_rosa   = PatternFill('solid', fgColor='F2CEEF')
fill_laranja= PatternFill('solid', fgColor='E97132')
fill_azul   = PatternFill('solid', fgColor='467886')
fill_amarelo= PatternFill('solid', fgColor='FFD700')

linhas_resumo = [
    ('RESUMO DA CONCILIAÇÃO', '', None),
    (f'Rodagem em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', '', None),
    ('', '', None),
    ('Categoria', 'Qtde. linhas', None),
    ('Remessas ROSA (100% zerado)', cnt_rem['rosa'], fill_rosa),
    ('Remessas LARANJA (parcial/incompleto)', cnt_rem['laranja'], fill_laranja),
    ('Remessas AMARELO (retorno c/ ICMS divergente)', cnt_rem['amarelo'], fill_amarelo),
    ('Retornos ROSA (ZEROU)', cnt_ret['rosa'], fill_rosa),
    ('Retornos LARANJA (PARCIAL)', cnt_ret['laranja'], fill_laranja),
    ('Retornos AZUL (sem remessa)', cnt_ret['azul'], fill_azul),
    ('Retornos AMARELO (ICMS do XML != lançado no SAP)', cnt_ret['amarelo'], fill_amarelo),
    ('', '', None),
    ('TOTAL DE LINHAS PINTADAS', len(pintura_p6), None),
    ('', '', None),
    ('Diagnóstico monetário', '', None),
    ('Saldo líquido ROSA (deve ser ~R$0)', round(soma_rosa_rem + soma_rosa_ret, 2), None),
    ('Saldo líquido LARANJA (em aberto)', round(soma_laranja_rem + soma_laranja_ret, 2), None),
    ('Saldo líquido AMARELO (revisar lançamento)', round(soma_amarelo_rem + soma_amarelo_ret, 2), None),
]

for i, (label, valor, fill) in enumerate(linhas_resumo, start=1):
    c1 = ws_resumo.cell(row=i, column=1, value=label)
    c2 = ws_resumo.cell(row=i, column=2, value=valor if valor != '' else None)
    if fill:
        c1.fill = fill; c2.fill = fill
    if label in ('RESUMO DA CONCILIAÇÃO',):
        c1.font = Font(bold=True, size=16)
    if label == 'Categoria':
        c1.font = font_header; c2.font = font_header
        c1.fill = fill_header; c2.fill = fill_header
    if label == 'TOTAL DE LINHAS PINTADAS':
        c1.font = Font(bold=True); c2.font = Font(bold=True)
    if isinstance(valor, float):
        c2.number_format = '#,##0.00'

ws_resumo.column_dimensions['A'].width = 42
ws_resumo.column_dimensions['B'].width = 16

# ---------------------------------------------------------------
# ABA "RELATORIO LARANJA" — lista so os casos que precisam de revisao
# (PARCIAL, SEM_REMESSA, SEM_RETORNO_P6), ordenados pela diferenca
# em modulo (do maior problema para o menor) — facilita revisar depois.
# ---------------------------------------------------------------
NOME_ABA_LARANJA = 'Relatorio Laranja'
if NOME_ABA_LARANJA in wb.sheetnames:
    del wb[NOME_ABA_LARANJA]
ws_laranja = wb.create_sheet(NOME_ABA_LARANJA, 1)  # logo depois do Resumo

casos_revisao = [r for r in relatorio_rows if r['status'] != 'ZEROU']
casos_revisao.sort(key=lambda r: abs(r['diferenca']), reverse=True)

cabecalho_laranja = [
    'NF Retorno', 'Status', 'Critério', 'Qtde. Remessas',
    'Valor Retorno', 'Valor Remessas Encontradas', 'Diferença',
    'vICMS no XML', 'Diferença ICMS (XML - lançado)',
    'Remessas Referenciadas',
]
for col, titulo in enumerate(cabecalho_laranja, start=1):
    c = ws_laranja.cell(row=1, column=col, value=titulo)
    c.font = font_header
    c.fill = fill_header

fill_atencao  = PatternFill('solid', fgColor='FFCCCC')   # diferenca grande generica
fill_div_icms = PatternFill('solid', fgColor='FFD700')   # divergencia de ICMS especificamente

for i, r in enumerate(casos_revisao, start=2):
    ws_laranja.cell(row=i, column=1, value=r['nnf'])
    ws_laranja.cell(row=i, column=2, value=r['status'])
    ws_laranja.cell(row=i, column=3, value=r['criterio'])
    ws_laranja.cell(row=i, column=4, value=r['n_remessas'])
    c5 = ws_laranja.cell(row=i, column=5, value=round(r['val_retorno'], 2))
    c6 = ws_laranja.cell(row=i, column=6, value=round(r['val_remessas'], 2))
    c7 = ws_laranja.cell(row=i, column=7, value=round(r['diferenca'], 2))
    c8 = ws_laranja.cell(row=i, column=8, value=round(r['vicms_xml'], 2) if r['vicms_xml'] is not None else None)
    c9 = ws_laranja.cell(row=i, column=9, value=r['dif_icms'])
    for c in (c5, c6, c7, c8, c9):
        c.number_format = '#,##0.00'
    ws_laranja.cell(row=i, column=10, value=r['referencias'])

    # Destaca: amarelo para divergencia de ICMS (o caso mais critico),
    # vermelho claro para outras diferencas grandes (> R$1.000)
    if r['status'] == 'DIVERGENTE_ICMS':
        fill_linha = fill_div_icms
    elif abs(r['diferenca']) > 1000:
        fill_linha = fill_atencao
    else:
        fill_linha = None
    if fill_linha:
        for col in range(1, 11):
            ws_laranja.cell(row=i, column=col).fill = fill_linha

for col, largura in zip('ABCDEFGHIJ', [12, 18, 22, 14, 16, 22, 16, 16, 22, 40]):
    ws_laranja.column_dimensions[col].width = largura
ws_laranja.freeze_panes = 'A2'

print(f"Aba '{NOME_ABA_LARANJA}' criada com {len(casos_revisao)} casos para revisao "
      f"(ordenados por diferenca, maiores primeiro).")

# Gera nome de saida automaticamente, sem sobrescrever rodagens anteriores
# e sem travar se uma versao anterior estiver aberta no Excel
base_output = ARQUIVO.replace('.xlsx', '_CONCILIADO')
n = 1
while True:
    candidato = f"{base_output}.xlsx" if n == 1 else f"{base_output}_v{n}.xlsx"
    if not os.path.exists(candidato):
        OUTPUT = candidato
        break
    n += 1

wb.save(OUTPUT)
print(f"Arquivo salvo: {OUTPUT}")
print(f"Aba '{NOME_ABA_RESUMO}' criada com o resumo desta rodagem.")

# ---------------------------------------------------------------
# RESUMO NO CONSOLE
# ---------------------------------------------------------------
print()
print('=' * 75)
print('RESULTADO DA CONCILIACAO')
print('=' * 75)
for linha in log_result:
    print(linha)

print()
print('RESUMO DOS RETORNOS (status da conciliacao):')
print(f'  Zerou   (rosa):              {cnt["ZEROU"]}')
print(f'  Parcial (laranja):           {cnt["PARCIAL"]}')
print(f'  Sem remessa (azul):          {cnt["SEM_REMESSA"]}')
print(f'  Divergente ICMS (amarelo):   {cnt.get("DIVERGENTE_ICMS", 0)}  <-- XML != valor lancado no SAP')
print(f'  Sem retorno P6:              {cnt["SEM_RAZAO"]}')
print()
print('RESUMO DAS REMESSAS PINTADAS:')
print(f'  Rosa    (100% zerado por retornos ZEROU): {cnt_rem["rosa"]}')
print(f'  Laranja (parcial ou retorno incompleto):  {cnt_rem["laranja"]}')
print(f'  Amarelo (retorno c/ ICMS divergente):     {cnt_rem["amarelo"]}')
print()
print(f'  Total linhas pintadas: {len(pintura_p6)}')
print()
print('Regra aplicada:')
print('  REMESSA fica ROSA apenas se:')
print('    1. Todo seu valor foi consumido (saldo = 0)')
print('    2. TODOS os retornos que a referenciam sao ZEROU')
print('    3. O ICMS do XML do(s) retorno(s) bate com o valor lancado no SAP')
print('  Caso contrario: LARANJA (incompleto) ou AMARELO (ICMS divergente, revisar lancamento)')
print()
print('Filtrar pela cor ROSA na Planilha6 deve resultar em soma = 0.')
print()

# Diagnostico monetario: soma por cor e tipo (ja calculado mais acima, reaproveitado aqui)
print('DIAGNOSTICO MONETARIO (valores das linhas pintadas):')
print(f'  ROSA   remessas (positivo): R$ {soma_rosa_rem:>15,.2f}')
print(f'  ROSA   retornos (negativo): R$ {soma_rosa_ret:>15,.2f}')
print(f'  ROSA   saldo liquido:        R$ {(soma_rosa_rem + soma_rosa_ret):>15,.2f}  <-- deve ser ~0')
print()
print(f'  LARANJA remessas:            R$ {soma_laranja_rem:>15,.2f}')
print(f'  LARANJA retornos:            R$ {soma_laranja_ret:>15,.2f}')
print(f'  LARANJA saldo liquido:       R$ {(soma_laranja_rem + soma_laranja_ret):>15,.2f}')
print()
print(f'  AMARELO remessas:            R$ {soma_amarelo_rem:>15,.2f}')
print(f'  AMARELO retornos:            R$ {soma_amarelo_ret:>15,.2f}')
print(f'  AMARELO saldo liquido:       R$ {(soma_amarelo_rem + soma_amarelo_ret):>15,.2f}  <-- revisar lancamento')
print()
saldo_total = soma_rosa_rem + soma_rosa_ret + soma_laranja_rem + soma_laranja_ret + soma_amarelo_rem + soma_amarelo_ret
print(f'  TOTAL pintado (rem+ret):     R$ {saldo_total:>15,.2f}')
if abs(soma_rosa_rem + soma_rosa_ret) < 1.0:
    print('  [OK] Rosa balanceado — filtro rosa soma ~R$0')
else:
    print(f'  [ATENCAO] Rosa fora de balanco por R$ {soma_rosa_rem + soma_rosa_ret:,.2f}')
